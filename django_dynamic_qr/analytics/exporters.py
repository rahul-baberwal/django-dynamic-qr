"""
Export scan logs to CSV or XLSX.
"""
import csv
import io
from django.http import HttpResponse


HEADERS = ["scanned_at", "country", "city", "device", "os", "browser", "ip_address", "referer"]


def _get_rows(qr_code):
    return qr_code.scans.values_list(
        "scanned_at", "country", "city", "device", "os", "browser", "ip_address", "referer"
    ).order_by("-scanned_at")


def export_scan_csv(qr_code) -> HttpResponse:
    response = HttpResponse(content_type="text/csv")
    response["Content-Disposition"] = f'attachment; filename="scans-{qr_code.slug}.csv"'
    writer = csv.writer(response)
    writer.writerow(HEADERS)
    for row in _get_rows(qr_code):
        writer.writerow(row)
    return response


def export_scan_xlsx(qr_code) -> HttpResponse:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Scan Logs"

    # Header row
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="4F46E5")
    for col, header in enumerate(HEADERS, 1):
        cell = ws.cell(row=1, column=col, value=header.replace("_", " ").title())
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    for row_num, row in enumerate(_get_rows(qr_code), 2):
        for col, value in enumerate(row, 1):
            ws.cell(row=row_num, column=col, value=str(value) if value else "")

    for col in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    response = HttpResponse(
        buf.read(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="scans-{qr_code.slug}.xlsx"'
    return response
